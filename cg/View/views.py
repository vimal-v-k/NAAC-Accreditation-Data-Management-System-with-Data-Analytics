from django.shortcuts import render,redirect
from .models import Details3_1_1,Details3_2_2,Details3_3_1,Details3_3_2,Details3_4_3,Details3_5_1,Citations
from django.http import HttpResponse
from django.contrib import messages
from Login.views import log
from openpyxl import Workbook


def vieww(request):
    return render(request,'View/view.html')
def viewmain(request):
    return render(request,'View/viewmain.html')
def viewdetails(request):
    return render(request,'View/viewdetails.html')



def logout_vw(request):
    try:
        del request.session['user_id']
    except KeyError:
        pass
    return redirect(log)



def f3_1_1(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Name_of_the_research_project_endowment = request.POST.get('Name of the research project/ endowment')
        Name_of_the_Principal_Investigator_Co_investigator = request.POST.get('Name of the Principal Investigator/Co-investigator')
        Department_of_Principal_Investigator = request.POST.get('Department of Principal Investigator')
        Year_of_Award = request.POST.get('Year of Award')
        Amount_Sanctioned = request.POST.get('Amount Sanctioned')
        Duration_of_the_project = request.POST.get('Duration of the project')
        Name_of_the_Funding_Agency = request.POST.get('Name of the Funding Agency')
        radiobutton3_1_1 = request.POST.get('radiobutton3_1_1')
        print(user_email,Name_of_the_research_project_endowment,Name_of_the_Principal_Investigator_Co_investigator,Department_of_Principal_Investigator,Year_of_Award,Amount_Sanctioned,Duration_of_the_project,Name_of_the_Funding_Agency,radiobutton3_1_1)

        if Name_of_the_research_project_endowment and Name_of_the_Principal_Investigator_Co_investigator and Department_of_Principal_Investigator and Year_of_Award and Amount_Sanctioned and Duration_of_the_project and Name_of_the_Funding_Agency and radiobutton3_1_1:
            if Details3_1_1.objects.filter(
                    Name_of_the_research_project_endowment=Name_of_the_research_project_endowment,
                    Name_of_the_Principal_Investigator_Co_investigator=Name_of_the_Principal_Investigator_Co_investigator,
                    Department_of_Principal_Investigator = Department_of_Principal_Investigator,
                    Year_of_Award = Year_of_Award,
                    Amount_Sanctioned = Amount_Sanctioned,
                    Duration_of_the_project =Duration_of_the_project,
                    Name_of_the_Funding_Agency = Name_of_the_Funding_Agency,
                    Government_non_Government = radiobutton3_1_1).exists():
                # If a record exists, display an error message and redirect the user back to the same page
                return redirect(f3_1_1)
            else:
                post = Details3_1_1()
                post.Uploaded_by = user_email
                post.Name_of_the_research_project_endowment =  Name_of_the_research_project_endowment
                post.Name_of_the_Principal_Investigator_Co_investigator = Name_of_the_Principal_Investigator_Co_investigator
                post.Department_of_Principal_Investigator = Department_of_Principal_Investigator
                post.Year_of_Award = Year_of_Award
                post.Amount_Sanctioned = Amount_Sanctioned
                post.Duration_of_the_project = Duration_of_the_project
                post.Name_of_the_Funding_Agency = Name_of_the_Funding_Agency
                post.Government_non_Government = radiobutton3_1_1
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.error(request,'Please fill all fields.')
            return redirect(f3_1_1)
    else:
        return render(request, 'View/3._1_1form.html')

def f3_2_2(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Year = request.POST.get('Year')
        Name_of_the_workshop_seminar_conference = request.POST.get('Name_of_the_workshop_seminar_conference')
        Number_of_Participants = request.POST.get('Number_of_Participants')
        Date_From = request.POST.get('Date_From')
        Date_To = request.POST.get('Date_To')
        radiobutton322 = request.POST.get('radiobutton322')
        Link_to_the_Activity_report_on_the_website = request.POST.get('Link_to_the_Activity_report_on_the_website')
        # print(user_email ,Year,Name_of_the_workshop_seminar_conference,Number_of_Participants,Date_From,Date_To,Link_to_the_Activity_report_on_the_website)

        if Year and Name_of_the_workshop_seminar_conference and Number_of_Participants and Date_From and Date_To and Link_to_the_Activity_report_on_the_website:
            if Details3_2_2.objects.filter(Year=Year,
                                           Name_of_the_workshop_seminar_conference=Name_of_the_workshop_seminar_conference,
                                           Number_of_Participants=Number_of_Participants, Date_From=Date_From,
                                           Date_To=Date_To,
                                            radiobutton322=radiobutton322,
                                           Link_to_the_Activity_report_on_the_website=Link_to_the_Activity_report_on_the_website).exists():
                return redirect(f3_2_2)
            else:
                post = Details3_2_2(
                    Year=Year,
                    Uploaded_by=user_email,
                    Name_of_the_workshop_seminar_conference=Name_of_the_workshop_seminar_conference,
                    Number_of_Participants=Number_of_Participants,
                    Date_From=Date_From,
                    Date_To=Date_To,
                    radiobutton322=radiobutton322,
                    Link_to_the_Activity_report_on_the_website=Link_to_the_Activity_report_on_the_website
                )
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.warning(request, 'Please fill all fields.')

            return redirect(f3_2_2)
    else:
        return render(request, 'View/3_2_2form.html')


def f3_3_1(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Title_of_paper = request.POST.get('Title_of_paper')
        Name_of_the_authors = request.POST.get('Name_of_the_author/s')
        radiobutton3_2_2 = request.POST.get('radiobutton3.2.2')
        Name_of_journal = request.POST.get('Name_of_journal')
        Year_of_publication = request.POST.get('Year_of_publication')
        ISSN_number = request.POST.get('ISSN_number')
        Link_to_website_of_the_Journal = request.POST.get('Link_to_website_of_the_Journal')
        Link_to_article_paper_abstract_of_the_article = request.POST.get('Link_to_article_paper_abstract_of_the_article')
        UGC_Care_list = request.POST.get("radiobutton3.2.2.1")
        # print(Title_of_paper,Name_of_the_authors,radiobutton3_2_2,Name_of_journal,Year_of_publication,ISSN_number, Link_to_website_of_the_Journal, Link_to_article_paper_abstract_of_the_article,UGC_Care_list)

        if Title_of_paper and Name_of_the_authors and radiobutton3_2_2 and Name_of_journal and Year_of_publication and ISSN_number and Link_to_website_of_the_Journal and Link_to_article_paper_abstract_of_the_article and UGC_Care_list:
            existing_data = Details3_3_1.objects.filter(
                Title_of_paper=Title_of_paper,
                Name_of_the_authors=Name_of_the_authors,
                Department_of_the_teacher=radiobutton3_2_2,
                Name_of_journal=Name_of_journal,
                Year_of_publication=Year_of_publication,
                ISSN_number=ISSN_number,
                Link_to_website_of_the_Journal=Link_to_website_of_the_Journal,
                Link_to_article_paper_abstract_of_the_article=Link_to_article_paper_abstract_of_the_article,
                Is_it_listed_in_UGC_Care_list=UGC_Care_list
            )

            if existing_data:
                messages.warning(request, 'Data already exists in the database.')
                return redirect(f3_3_1)
            else:
                post = Details3_3_1()
                post.Uploaded_by = user_email
                post.Title_of_paper =  Title_of_paper
                post.Name_of_the_authors = Name_of_the_authors
                post.Department_of_the_teacher = radiobutton3_2_2
                post.Name_of_journal = Name_of_journal
                post.Year_of_publication = Year_of_publication
                post.ISSN_number = ISSN_number
                post.Link_to_website_of_the_Journal = Link_to_website_of_the_Journal
                post.Link_to_article_paper_abstract_of_the_article = Link_to_article_paper_abstract_of_the_article
                post.Is_it_listed_in_UGC_Care_list = UGC_Care_list
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.warning(request, 'Please fill all fields.')

            return redirect(f3_3_1)

    else:
        return render(request, 'View/3_3_1form.html')


def f3_3_2(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Name_of_the_teacher = request.POST.get('Name_of_the_teacher')
        Title_of_the_book_chapters_published = request.POST.get('Title_of_the_book_chapters_published')
        Title_of_the_paper = request.POST.get('Title_of_the_paper')
        Title_of_the_proceedings_of_the_conference = request.POST.get('Title_of_the_proceedings_of_the_conference')
        Name_of_the_conference = request.POST.get('Name_of_the_conference')
        radiobutton3_3_2 = request.POST.get('radiobutton3.3.2')
        Year_of_publications = request.POST.get('Year_of_publication')
        ISBN_number_of_the_proceeding = request.POST.get('ISBN_number_of_the_proceeding')
        Affiliating_Institute_at_the_time_of_publication = request.POST.get("Affiliating_Institute_at_the_time_of_publication")
        Name_of_the_publisher = request.POST.get("Name_of_the_publisher")
        # print(Name_of_the_teacher, Title_of_the_book_chapters_published, Title_of_the_paper, Title_of_the_proceedings_of_the_conference, Year_of_publications, Name_of_the_conference,radiobutton3_3_2, ISBN_number_of_the_proceeding, Affiliating_Institute_at_the_time_of_publication,Name_of_the_publisher)

        if Name_of_the_teacher and Title_of_the_book_chapters_published and Title_of_the_paper and Title_of_the_proceedings_of_the_conference and Year_of_publications and Name_of_the_conference and radiobutton3_3_2 and ISBN_number_of_the_proceeding and Affiliating_Institute_at_the_time_of_publication and Name_of_the_publisher:
            existing_data = Details3_3_2.objects.filter(
                Name_of_the_teacher=Name_of_the_teacher,
                Title_of_the_book_chapters_published=Title_of_the_book_chapters_published,
                Title_of_the_paper=Title_of_the_paper,
                Title_of_the_proceedings_of_the_conference=Title_of_the_proceedings_of_the_conference,
                Name_of_the_conference=Name_of_the_conference,
                national_international=radiobutton3_3_2,
                Year_of_publications=Year_of_publications,
                ISBN_number_of_the_proceeding=ISBN_number_of_the_proceeding,
                Affiliating_Institute_at_the_time_of_publication=Affiliating_Institute_at_the_time_of_publication,
                Name_of_the_publisher=Name_of_the_publisher,
            ).exists()

            if existing_data:
                messages.error(request, 'Data already exists.')
                return redirect(f3_3_2)
            else:
                post = Details3_3_2()
                post.Uploaded_by = user_email
                post.Name_of_the_teacher = Name_of_the_teacher
                post.Title_of_the_book_chapters_published = Title_of_the_book_chapters_published
                post.Title_of_the_paper = Title_of_the_paper
                post.Title_of_the_proceedings_of_the_conference = Title_of_the_proceedings_of_the_conference
                post.Name_of_the_conference = Name_of_the_conference
                post.national_international = radiobutton3_3_2
                post.Year_of_publications = Year_of_publications
                post.ISBN_number_of_the_proceeding = ISBN_number_of_the_proceeding
                post.Affiliating_Institute_at_the_time_of_publication = Affiliating_Institute_at_the_time_of_publication
                post.Name_of_the_publisher=Name_of_the_publisher
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.warning(request, 'Please fill all fields.')

            return redirect(f3_3_2)
    else:
        return render(request, 'View/3_3_2form.html')


def f3_4_3(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Name_of_the_activity = request.POST.get('Name_of_the_activity')
        Organising_unit_agency_collaborating_agency = request.POST.get('Organising_unit_agency_collaborating_agency')
        Name_of_the_scheme = request.POST.get('Name_of_the_scheme')
        Year_of_the_activity = request.POST.get('Year_of_the_activity')

        # print(Name_of_the_activity,Organising_unit_agency_collaborating_agency,Name_of_the_scheme,Year_of_the_activity)

        if Name_of_the_activity and Organising_unit_agency_collaborating_agency and Name_of_the_scheme and Year_of_the_activity:
            existing_entry = Details3_4_3.objects.filter(Name_of_the_activity=Name_of_the_activity,
                                                         Organising_unit_agency_collaborating_agency=Organising_unit_agency_collaborating_agency,
                                                         Name_of_the_scheme=Name_of_the_scheme,
                                                         Year_of_the_activity=Year_of_the_activity).exists()

            if existing_entry:
                messages.error(request,'This entry already exists.')
                return redirect(f3_4_3)

            else:
                post = Details3_4_3()
                post.Uploaded_by = user_email
                post.Name_of_the_activity = Name_of_the_activity
                post.Organising_unit_agency_collaborating_agency = Organising_unit_agency_collaborating_agency
                post.Name_of_the_scheme = Name_of_the_scheme
                post.Year_of_the_activity = Year_of_the_activity
                print(post.Name_of_the_activity, post.Organising_unit_agency_collaborating_agency,post.Name_of_the_scheme,post.Year_of_the_activity)
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.warning(request, 'Please fill all fields.')

            return redirect(f3_4_3)#
    else:
        return render(request, 'View/3.4.3.html')

def f3_5_1(request):
    if request.method == "POST":
        user_email = request.session.get('user_email', None)
        Name_of_the_MoU_Collaboration_linkage = request.POST.get('Name_of_the_MoU_Collaboration_linkage')
        Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details = request.POST.get('Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details')
        Year_of_signing_MoU_collaboration_linkage = request.POST.get('Year_of_signing_MoU_collaboration_linkage')
        radiobutton3_5_1 = request.POST.get('radiobutton3_5_1')
        List_the_actual_activities_under_each_MOU_and_web_inks_year_wise = request.POST.get('List_the_actual_activities_under_each_MOU_and_web_inks_year_wise')
        Link_to_the_relevant_document = request.POST.get('Link_to_the_relevant_document')

        # print(Name_of_the_MoU_Collaboration_linkage,Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details,Year_of_signing_MoU_collaboration_linkage,List_the_actual_activities_under_each_MOU_and_web_inks_year_wise,Link_to_the_relevant_document)

        if  Name_of_the_MoU_Collaboration_linkage and radiobutton3_5_1 and Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details and Year_of_signing_MoU_collaboration_linkage and List_the_actual_activities_under_each_MOU_and_web_inks_year_wise and Link_to_the_relevant_document :
            if Details3_5_1.objects.filter(Name_of_the_MoU_Collaboration_linkage=Name_of_the_MoU_Collaboration_linkage,
                                           Year_of_signing_MoU_collaboration_linkage=Year_of_signing_MoU_collaboration_linkage,Link_to_the_relevant_document = Link_to_the_relevant_document,
                                           List_the_actual_activities_under_each_MOU_and_web_inks_year_wise = List_the_actual_activities_under_each_MOU_and_web_inks_year_wise,
                                           radiobutton3_5_1 = radiobutton3_5_1,
                                           Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details = Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details).exists():
                messages.error(request, 'Data already exists.')
                return redirect(f3_5_1)
            else:
                post = Details3_5_1()
                post.Uploaded_by = user_email
                post.Name_of_the_MoU_Collaboration_linkage = Name_of_the_MoU_Collaboration_linkage
                post.Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details = Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details
                post.Year_of_signing_MoU_collaboration_linkage = Year_of_signing_MoU_collaboration_linkage
                post.radiobutton3_5_1=radiobutton3_5_1
                post.List_the_actual_activities_under_each_MOU_and_web_inks_year_wise = List_the_actual_activities_under_each_MOU_and_web_inks_year_wise
                post.Link_to_the_relevant_document = Link_to_the_relevant_document

                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.warning(request, 'Please fill all fields.')

            return redirect(f3_5_1)
    else:
        return render(request, 'View/3.5.1.html')


def viewdetails(request):
    return render(request,'View/viewdetails.html')


def view311(request):
    user_email = request.session.get('user_email', None)
    data = Details3_1_1.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/311details.html', {'data': data})
def export_details3_1_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details3_1_1.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Name of the research project/ endowment',
               'Name of the Principal Investigator/Co-investigator',
               'Department of Principal Investigator',
               'Year of Award',
               'Amount Sanctioned',
               'Duration of the project',
               'Name of the Funding Agency',
               'Gov/Non-Gov']
    row_num = 1

    # Write the column headers to the Excel file
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Query the database for all instances of the Details3_5_1 model
    queryset = Details3_1_1.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_research_project_endowment,
               obj.Name_of_the_Principal_Investigator_Co_investigator,
               obj.Department_of_Principal_Investigator,
               obj.Year_of_Award,
               obj.Amount_Sanctioned,
               obj.Duration_of_the_project,
               obj.Name_of_the_Funding_Agency,
               obj.Government_non_Government]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    print()
    return response
    # return render(response, 'View/view351.html')


def view322(request):
    user_email = request.session.get('user_email', None)
    data = Details3_2_2.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/322details.html', {'data': data})
def export_details3_2_2(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details3_2_2.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Year',
               'Name of the workshop/ seminar/ conference',
               'Number of Participants',
               'Date From',
               'Date To',
               'Link to the Activity report on the website']
    row_num = 1

    # Write the column headers to the Excel file
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Query the database for all instances of the Details3_2_2 model
    queryset = Details3_2_2.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Year,
               obj.Name_of_the_workshop_seminar_conference,
               obj.Number_of_Participants,
               obj.Date_From,
               obj.Date_To,
               obj.Link_to_the_Activity_report_on_the_website]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    print()
    return response



def view331(request):
    user_email = request.session.get('user_email', None)
    data = Details3_3_1.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/331details.html', {'data': data})
def export_details3_3_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details331.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Title of paper',
               'Name of the author/s',
               'Department of the teacher',
               'Name of journal',
               'Year of publication',
               'ISSN number',
               'Link to website of the Journal',
               'Link to article / paper / abstract of the article',
               'Is it listed in UGC Care list']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_3_1.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Title_of_paper,
               obj.Name_of_the_authors,
               obj.Department_of_the_teacher,
               obj.Name_of_journal,
               obj.Year_of_publication,
               obj.ISSN_number,
               obj.Link_to_website_of_the_Journal,
               obj.Link_to_article_paper_abstract_of_the_article,
               obj.Is_it_listed_in_UGC_Care_list]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response



def view332(request):
    user_email = request.session.get('user_email', None)
    data = Details3_3_2.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/332details.html', {'data': data})
def export_details3_3_2(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details332.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Name of the teacher',
               'Title of the book/chapters  published',
               'Title of the paper',
               'Title of the proceedings of the conference',
               'Name of the conference',
               'national/international',
               'Year of publication',
               'IISBN number of the proceeding',
               'Affiliating Institute at the time of publication',
               'Name of the publisher']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_3_2.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_teacher,
               obj.Title_of_the_book_chapters_published,
               obj.Title_of_the_paper,
               obj.Title_of_the_proceedings_of_the_conference,
               obj.Name_of_the_conference,
               obj.national_international,
               obj.Year_of_publication,
               obj.ISBN_number_of_the_proceeding,
               obj.Affiliating_Institute_at_the_time_of_publication,
               obj.Name_of_the_publisher]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response



def view343(request):
    user_email = request.session.get('user_email', None)
    data = Details3_4_3.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/343details.html', {'data': data})
def export_details3_4_3(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details343.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Name of the activity',
               'Organising unit/ agency/ collaborating agency',
               'Name of the scheme',
               'Year of the activity']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_4_3.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_activity,
               obj.Organising_unit_agency_collaborating_agency,
               obj.Name_of_the_scheme,
               obj.Year_of_the_activity]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    print()
    return response




def view351(request):
    user_email = request.session.get('user_email', None)
    data = Details3_5_1.objects.filter(Uploaded_by=user_email)
    return render(request, 'View/351details.html', {'data': data})
def export_details3_5_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details351.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)

    # Define the column headers for the Excel file
    columns = ['Name of the MoU / Collaboration / linkage',
               'Name of the collaborating agency / institution / industry / corporate house with whom the MoU / collaboration / linkage is made, with contact details',
               'Year of signing MoU / collaboration / linkage',
               'Duration of MoU / collaboration / linkage',
               'List the  actual  activities under each MOU  and web -links year-wise',
               'Link to the relevant document']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_5_1.objects.filter(Uploaded_by=user_email)

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_MoU_Collaboration_linkage,
               obj.Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details,
               obj.Year_of_signing_MoU_collaboration_linkage,
               obj.radiobutton3_5_1,
               obj.List_the_actual_activities_under_each_MOU_and_web_inks_year_wise,
               obj.Link_to_the_relevant_document]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response


def edit311(request, id):
        d311 = Details3_1_1.objects.get(id=id)
        if request.method == "POST":
            d311.Name_of_the_research_project_endowment = request.POST.get('Name of the research project/ endowment')
            d311.Name_of_the_Principal_Investigator_Co_investigator = request.POST.get('Name of the Principal Investigator/Co-investigator')
            d311.Department_of_Principal_Investigator = request.POST.get('Department of Principal Investigator')
            d311.Year_of_Award = request.POST.get('Year of Award')
            d311.Amount_Sanctioned = request.POST.get('Amount Sanctioned')
            d311.Duration_of_the_project = request.POST.get('Duration of the project')
            d311.Name_of_the_Funding_Agency = request.POST.get('Name of the Funding Agency')
            d311.Government_non_Government = request.POST.get('radiobutton3.1.1')
            d311.save()
            messages.warning(request, 'Updated Successfully')

            return redirect(view311)
        return render(request,'View/311edit.html',{'d311' : d311})

def edit322(request, id):
    d322 = Details3_2_2.objects.get(id=id)
    if request.method == "POST":
        d322.Year = request.POST.get('Year')
        d322.Name_of_the_workshop_seminar_conference = request.POST.get('Name_of_the_workshop_seminar_conference')
        d322.Number_of_Participants = request.POST.get('Number_of_Participants')
        d322. Date_From= request.POST.get('Date_From')
        d322.Date_To = request.POST.get('Date_To')
        d322.Link_to_the_Activity_report_on_the_website = request.POST.get('Link_to_the_Activity_report_on_the_website')
        d322.save()
        messages.warning(request, 'Updated Successfully')

        return redirect(view322)

    return render(request,'View/322edit.html',{'d322' : d322})


def edit331(request, id):
    d331 = Details3_3_1.objects.get(id=id)
    if request.method == "POST":
        d331.Year = request.POST.get('Year')
        d331.Title_of_paper = request.POST.get('Title_of_paper')
        d331.Name_of_the_authors = request.POST.get('Name_of_the_author/s')
        d331.radiobutton3_2_2 = request.POST.get('radiobutton3.2.2')
        d331.Name_of_journal = request.POST.get('Name_of_journal')
        d331.Year_of_publication = request.POST.get('Year_of_publication')
        d331.ISSN_number = request.POST.get('ISSN_number')
        d331.Link_to_website_of_the_Journal = request.POST.get('Link_to_website_of_the_Journal')
        d331.Link_to_article_paper_abstract_of_the_article = request.POST.get('Link_to_article_paper_abstract_of_the_article')
        d331.UGC_Care_list = request.POST.get("radiobutton3.2.2.1")
        d331.save()
        messages.warning(request, 'Updated Successfully')

        return redirect(view331)

    return render(request,'View/331edit.html',{'d331' : d331})


def edit332(request, id):
    d332 = Details3_3_2.objects.get(id=id)
    if request.method == "POST":
        d332.Name_of_the_teacher = request.POST.get('Name_of_the_teacher')
        d332.Title_of_the_book_chapters_published = request.POST.get('Title_of_the_book_chapters_published')
        d332.Title_of_the_paper = request.POST.get('Title_of_the_paper')
        d332.Title_of_the_proceedings_of_the_conference = request.POST.get('Title_of_the_proceedings_of_the_conference')
        d332.Name_of_the_conference = request.POST.get('Name_of_the_conference')
        d332.radiobutton3_3_2 = request.POST.get('radiobutton3.3.2')
        d332.Year_of_publications = request.POST.get('Year_of_publication')
        d332.ISBN_number_of_the_proceeding = request.POST.get('ISBN_number_of_the_proceeding')
        d332.Affiliating_Institute_at_the_time_of_publication = request.POST.get("Affiliating_Institute_at_the_time_of_publication")
        d332.Name_of_the_publisher = request.POST.get("Name_of_the_publisher")
        d332.save()
        messages.warning(request, 'Updated Successfully')

        return redirect(view332)

    return render(request,'View/332edit.html',{'d332' : d332})


def edit343(request, id):
    d343 = Details3_4_3.objects.get(id=id)
    if request.method == "POST":
        d343.Name_of_the_activity = request.POST.get('Name_of_the_activity')
        d343.Organising_unit_agency_collaborating_agency = request.POST.get('Organising_unit_agency_collaborating_agency')
        d343.Name_of_the_scheme = request.POST.get('Name_of_the_scheme')
        d343.Year_of_the_activity = request.POST.get('Year_of_the_activity')
        d343.save()
        messages.warning(request, 'Updated Successfully')

        return redirect(view343)

    return render(request,'View/343edit.html',{'d343': d343})


def edit351(request,id):
    d351=Details3_5_1.objects.get(id=id)
    # print(d351.Year_of_signing_MoU_collaboration_linkage)
    if request.method == "POST":
        d351.Name_of_the_MoU_Collaboration_linkage= request.POST.get('Name_of_the_MoU_Collaboration_linkage')
        d351.Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details= request.POST.get('Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details')
        d351.Year_of_signing_MoU_collaboration_linkage= request.POST.get('Year_of_signing_MoU_collaboration_linkage')
        d351.radiobutton3_5_1= request.POST.get('radiobutton3_5_1')
        d351.List_the_actual_activities_under_each_MOU_and_web_inks_year_wise= request.POST.get('List_the_actual_activities_under_each_MOU_and_web_inks_year_wise')
        d351.Link_to_the_relevant_document= request.POST.get('Link_to_the_relevant_document')
        d351.save()

        messages.warning(request, 'Updated Successfully')
        return redirect(view351)

    return render(request,'View/351edit.html',{'d351':d351})


def del311(request,id):
    del311=Details3_1_1.objects.get(id=id)
    del311.delete()
    messages.success(request, 'Deleted Successfully')

    return redirect(view311)

def del322(request,id):
    del322=Details3_2_2.objects.get(id=id)
    del322.delete()
    messages.success(request, 'Deleted Successfully')

    return redirect(view322)

def del331(request,id):
    del331=Details3_3_1.objects.get(id=id)
    del331.delete()
    messages.success(request, 'Deleted Successfully')

    return redirect(view331)

def del332(request,id):
    del332=Details3_3_2.objects.get(id=id)
    del332.delete()
    messages.success(request, 'Deleted Successfully')

    return redirect(view332)

def del343(request,id):
    del343=Details3_4_3.objects.get(id=id)
    del343.delete()
    messages.success(request, 'Deleted Successfully')

    return redirect(view343)

def del351(request,id):
    del351=Details3_5_1.objects.get(id=id)
    del351.delete()
    messages.success(request, 'Deleted successfully.')
    return redirect(view351)



def search311(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date
        projects = Details3_1_1.objects.filter(Uploaded_by=user_email, Year_of_Award__range=[start_date, end_date])

        return render(request, 'View/311search.html', {'projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/311search.html')

def search322(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date
        projects = Details3_2_2.objects.filter(Uploaded_by=user_email, Year__range=[start_date, end_date])

        return render(request, 'View/322search.html', {'322projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/322search.html')

def search331(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date
        projects = Details3_3_1.objects.filter(Uploaded_by=user_email, Year_of_publication__range=[start_date, end_date])

        return render(request, 'View/331search.html', {'331projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/331search.html')

def search332(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date
        projects = Details3_3_2.objects.filter(Uploaded_by=user_email, Year_of_publications__range=[start_date, end_date])
        return render(request, 'View/332search.html', {'332projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/332search.html')
def search343(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date
        projects = Details3_4_3.objects.filter(Uploaded_by=user_email, Year_of_the_activity__range=[start_date, end_date])

        return render(request, 'View/343search.html', {'343projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/343search.html')

def search351(request):
    if request.method == 'POST':
        user_email = request.session.get('user_email', None )
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        if not start_date or not end_date:
            messages.warning(request, 'Please provide valid start and end dates to perform the search.')
            return redirect(view351)
        request.session['start_date'] = start_date
        request.session['end_date'] = end_date

        projects = Details3_5_1.objects.filter(Uploaded_by=user_email, Year_of_signing_MoU_collaboration_linkage__range=[start_date, end_date])

        return render(request, 'View/351search.html', {'351projects': projects,'start_date': start_date, 'end_date': end_date})
    else:
        return render(request, 'View/351search.html')

def gsearch311(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_1_1.objects.filter(Uploaded_by=user_email,Name_of_the_research_project_endowment__icontains=query)
    return render(request,'View/gsearch311.html',{'sg311': sgdata})

def gsearch322(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_2_2.objects.filter(Uploaded_by=user_email,Name_of_the_workshop_seminar_conference__icontains=query)
    return render(request,'View/gsearcg322.html',{'sg322': sgdata})

def gsearch331(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_3_1.objects.filter(Uploaded_by=user_email,Title_of_paper__icontains=query)
    return render(request,'View/gsearch331.html',{'sg331': sgdata})

def gsearch332(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_3_2.objects.filter(Uploaded_by=user_email,Name_of_the_conference__icontains=query)
    return render(request,'View/gsearch332.html',{'sg332': sgdata})

def gsearch343(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_4_3.objects.filter(Uploaded_by=user_email,Name_of_the_activity__icontains=query)
    return render(request,'View/gsearch343.html',{'sg343': sgdata})

def gsearch351(request):
    query = request.GET['search']
    user_email = request.session.get('user_email', None)
    sgdata = Details3_5_1.objects.filter(Uploaded_by=user_email,Name_of_the_MoU_Collaboration_linkage__icontains=query)
    return render(request,'View/gsearch351.html',{'sg351': sgdata})


def export_filterdetails3_5_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details351.xlsx"'
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # start_date = request.POST.get('start_date')
    # end_date = request.POST.get('end_date')

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)


    # Define the column headers for the Excel file
    columns = ['Name of the MoU / Collaboration / linkage',
               'Name of the collaborating agency / institution / industry / corporate house with whom the MoU / collaboration / linkage is made, with contact details',
               'Year of signing MoU / collaboration / linkage',
               'Duration of MoU / collaboration / linkage',
               'List the  actual  activities under each MOU  and web -links year-wise',
               'Link to the relevant document']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_5_1.objects.filter(Uploaded_by=user_email, Year_of_signing_MoU_collaboration_linkage__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_MoU_Collaboration_linkage,
               obj.Name_of_the_collaborating_agency_institution_industry_corporate_house_with_whom_the_MoU_collaboration_linkage_is_made_with_contact_details,
               obj.Year_of_signing_MoU_collaboration_linkage,
               obj.radiobutton3_5_1,
               obj.List_the_actual_activities_under_each_MOU_and_web_inks_year_wise,
               obj.Link_to_the_relevant_document]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response


def export_filterdetails3_4_3(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details343.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # Define the column headers for the Excel file
    columns = ['Name of the activity',
               'Organising unit/ agency/ collaborating agency',
               'Name of the scheme',
               'Year of the activity']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_4_3.objects.filter(Uploaded_by=user_email, Year_of_the_activity__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_activity,
               obj.Organising_unit_agency_collaborating_agency,
               obj.Name_of_the_scheme,
               obj.Year_of_the_activity]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response


def export_filterdetails3_3_2(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details332.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # Define the column headers for the Excel file
    columns = ['Name of the teacher',
               'Title of the book/chapters  published',
               'Title of the paper',
               'Title of the proceedings of the conference',
               'Name of the conference',
               'national/international',
               'Year of publication',
               'IISBN number of the proceeding',
               'Affiliating Institute at the time of publication',
               'Name of the publisher']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_3_2.objects.filter(Uploaded_by=user_email,Year_of_publications__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_teacher,
               obj.Title_of_the_book_chapters_published,
               obj.Title_of_the_paper,
               obj.Title_of_the_proceedings_of_the_conference,
               obj.Name_of_the_conference,
               obj.national_international,
               obj.Year_of_publication,
               obj.ISBN_number_of_the_proceeding,
               obj.Affiliating_Institute_at_the_time_of_publication,
               obj.Name_of_the_publisher]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response

def export_filterdetails3_1_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details3_1_1.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # Define the column headers for the Excel file
    columns = ['Name of the research project/ endowment',
               'Name of the Principal Investigator/Co-investigator',
               'Department of Principal Investigator',
               'Year of Award',
               'Amount Sanctioned',
               'Duration of the project',
               'Name of the Funding Agency',
               'Gov/Non-Gov']
    row_num = 1

    # Write the column headers to the Excel file
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Query the database for all instances of the Details3_5_1 model
    queryset = Details3_1_1.objects.filter(Uploaded_by=user_email,Year_of_Award__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Name_of_the_research_project_endowment,
               obj.Name_of_the_Principal_Investigator_Co_investigator,
               obj.Department_of_Principal_Investigator,
               obj.Year_of_Award,
               obj.Amount_Sanctioned,
               obj.Duration_of_the_project,
               obj.Name_of_the_Funding_Agency,
               obj.Government_non_Government]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response


def export_filterdetails3_2_2(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details3_2_2.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # Define the column headers for the Excel file
    columns = ['Year',
               'Name of the workshop/ seminar/ conference',
               'Number of Participants',
               'Date From',
               'Date To',
               'Link to the Activity report on the website']
    row_num = 1

    # Write the column headers to the Excel file
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Query the database for all instances of the Details3_2_2 model
    queryset = Details3_2_2.objects.filter(Uploaded_by=user_email,Year__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Year,
               obj.Name_of_the_workshop_seminar_conference,
               obj.Number_of_Participants,
               obj.Date_From,
               obj.Date_To,
               obj.Link_to_the_Activity_report_on_the_website]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response


def export_filterdetails3_3_1(request):
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="details331.xlsx"'

    # Create a new Excel workbook and select the active worksheet
    workbook = Workbook()
    worksheet = workbook.active
    user_email = request.session.get('user_email', None)
    start_date = request.session.get('start_date', None)
    end_date = request.session.get('end_date', None)

    # Define the column headers for the Excel file
    columns = ['Title of paper',
               'Name of the author/s',
               'Department of the teacher',
               'Name of journal',
               'Year of publication',
               'ISSN number',
               'Link to website of the Journal',
               'Link to article / paper / abstract of the article',
               'Is it listed in UGC Care list']
    row_num = 1

    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    queryset = Details3_3_1.objects.filter(Uploaded_by=user_email,Year_of_publication__range=[start_date, end_date])

    # Write the model data to the Excel file
    for obj in queryset:
        row_num += 1
        row = [obj.Title_of_paper,
               obj.Name_of_the_authors,
               obj.Department_of_the_teacher,
               obj.Name_of_journal,
               obj.Year_of_publication,
               obj.ISSN_number,
               obj.Link_to_website_of_the_Journal,
               obj.Link_to_article_paper_abstract_of_the_article,
               obj.Is_it_listed_in_UGC_Care_list]

        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    # Save the workbook and return the Excel file as an HttpResponse
    workbook.save(response)
    return response

def citations(request):
    if request.method == "POST":
        Name = request.POST.get('Namee')
        Email = request.POST.get('Email')
        user_email = request.session.get('user_email', None)
        citations = request.POST.get('citations')
        h_index = request.POST.get('h-index')
        i10_index = request.POST.get('i10-index')
        profile_link = request.POST.get('profile_link')
        Department =request.POST.get('radiobuttonc')
        year =request.POST.get('yearsc')



        if citations and h_index and i10_index and profile_link :
            if Citations.objects.filter(
                    Name=Name,Cemail=Email,Department=Department,year=year,
                    Citations=citations,
                    h_index=h_index,
                    i10_index = i10_index,
                    profile_link = profile_link).exists():
                # If a record exists, display an error message and redirect the user back to the same page
                return redirect(citations)
            else:
                post = Citations()
                post.Name=Name
                post.Cemail=Email
                post.Citations = citations
                post.h_index = h_index
                post.i10_index = i10_index
                post.profile_link = profile_link
                post.Department=Department
                post.year=year
                post.save()
                messages.success(request, 'Uploaded successfully.')

                return render(request, 'View/view.html')
        else:
            messages.error(request,'Please fill all fields.')
            return redirect(citations)
    return render(request,"View/Citations.html")